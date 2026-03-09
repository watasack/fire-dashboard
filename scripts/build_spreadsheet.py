"""
FIREシミュレーター Googleスプレッドシート生成スクリプト
出力: dist/fire_simulator.xlsx → Google Sheets にアップロードして使う

Usage:
    python scripts/build_spreadsheet.py
"""
import csv
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────
# ① 入力シートのセル行番号（全シートで共有）
# ──────────────────────────────────────────────
R = {}   # cell row numbers in ①入力

# ──────────────────────────────────────────────
# スタイル定数
# ──────────────────────────────────────────────
YELLOW  = PatternFill("solid", fgColor="FFFF99")
SKYBLUE = PatternFill("solid", fgColor="CCE5FF")
GREEN   = PatternFill("solid", fgColor="CCFFCC")
GRAY    = PatternFill("solid", fgColor="E0E0E0")
ORANGE  = PatternFill("solid", fgColor="FFD966")

BOLD    = Font(bold=True)
BOLD_LG = Font(bold=True, size=13)
BOLD_XL = Font(bold=True, size=16)
SMALL   = Font(size=9, color="666666")

def thin():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def cw(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

def section_header(ws, row, text):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, size=11)
    c.fill = ORANGE
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.row_dimensions[row].height = 18

def input_row(ws, row, label, default, unit, note=""):
    ws.cell(row=row, column=1, value=label)
    b = ws.cell(row=row, column=2, value=default)
    b.fill = YELLOW
    b.border = thin()
    b.alignment = Alignment(horizontal="right")
    ws.cell(row=row, column=3, value=unit)
    ws.cell(row=row, column=4, value=note).font = SMALL

def formula_row(ws, row, label, formula, unit, note=""):
    ws.cell(row=row, column=1, value=label)
    b = ws.cell(row=row, column=2, value=formula)
    b.fill = SKYBLUE
    b.border = thin()
    b.alignment = Alignment(horizontal="right")
    ws.cell(row=row, column=3, value=unit)
    ws.cell(row=row, column=4, value=note).font = SMALL


# ══════════════════════════════════════════════
# ① 入力シート
# ══════════════════════════════════════════════
def build_input_sheet(wb):
    ws = wb.create_sheet("①入力")
    cw(ws, "A", 32); cw(ws, "B", 18); cw(ws, "C", 10); cw(ws, "D", 36)

    # タイトル
    ws.row_dimensions[1].height = 32
    t = ws.cell(row=1, column=1, value="🔥 共働き・子育て世帯のFIREシミュレーター")
    t.font = BOLD_XL
    ws.merge_cells("A1:D1")
    sub = ws.cell(row=2, column=1, value="黄色セルに数値を入力 → 「③FIRE結果」シートで確認")
    sub.font = SMALL
    ws.merge_cells("A2:D2")

    r = 4   # 現在の行カウンター

    # ── 基本情報 ──
    section_header(ws, r, "👤  基本情報"); r += 1
    R["age_husband"] = r
    input_row(ws, r, "夫の現在年齢", 35, "歳"); r += 1
    R["age_wife"] = r
    input_row(ws, r, "妻の現在年齢", 33, "歳"); r += 1
    r += 1  # 空行

    # ── 収入 ──
    section_header(ws, r, "💰  収入（月手取り・税引後）"); r += 1
    R["income_husband"] = r
    input_row(ws, r, "夫の月手取り収入（通常勤務時）", 450000, "円/月",
              "育休・時短以外の通常時"); r += 1
    R["income_wife"] = r
    input_row(ws, r, "妻の月手取り収入（通常勤務時）", 300000, "円/月",
              "育休・時短以外の通常時"); r += 1
    R["income_total"] = r
    formula_row(ws, r, "世帯月手取り合計（自動計算）",
                f"=B{R['income_husband']}+B{R['income_wife']}", "円/月"); r += 1
    R["income_growth"] = r
    input_row(ws, r, "収入の年間成長率", 0.015, "%", "昇給・昇格等"); r += 1
    r += 1

    # ── 支出 ──
    section_header(ws, r, "🛒  支出"); r += 1
    R["expense_monthly"] = r
    input_row(ws, r, "月間生活費（現在）", 300000, "円/月", "住居費・保育料込み"); r += 1
    R["fire_expense_monthly"] = r
    input_row(ws, r, "FIRE後の月間生活費", 250000, "円/月", "子育て費用減少後を想定"); r += 1
    R["fire_expense_annual"] = r
    formula_row(ws, r, "FIRE後年間支出（自動計算）",
                f"=B{R['fire_expense_monthly']}*12", "円/年"); r += 1
    r += 1

    # ── 現在の資産 ──
    section_header(ws, r, "🏦  現在の資産"); r += 1
    R["asset_cash"] = r
    input_row(ws, r, "現金・預金", 3000000, "円", "普通・定期預金の合計"); r += 1
    R["asset_stocks"] = r
    input_row(ws, r, "株式・投資信託（NISA含む）", 10000000, "円"); r += 1
    R["asset_ideco"] = r
    input_row(ws, r, "iDeCo残高", 2000000, "円"); r += 1
    R["asset_total"] = r
    formula_row(ws, r, "合計資産（自動計算）",
                f"=B{R['asset_cash']}+B{R['asset_stocks']}+B{R['asset_ideco']}", "円"); r += 1
    r += 1

    # ── 育休・時短 ──
    section_header(ws, r, "👶  育休・時短設定（取得しない場合は 0）"); r += 1
    R["leave_husband_months"] = r
    input_row(ws, r, "夫の育休期間", 0, "ヶ月", "取得しない場合は 0"); r += 1
    R["leave_husband_income"] = r
    input_row(ws, r, "夫の育休中の月収（手当）", 100000, "円/月",
              "育休給付金 ≒ 給与の67%（上限あり）"); r += 1
    R["leave_wife_months"] = r
    input_row(ws, r, "妻の育休期間", 12, "ヶ月"); r += 1
    R["leave_wife_income"] = r
    input_row(ws, r, "妻の育休中の月収（手当）", 150000, "円/月"); r += 1
    R["part_wife_months"] = r
    input_row(ws, r, "妻の時短勤務期間（育休後）", 36, "ヶ月"); r += 1
    R["part_wife_income"] = r
    input_row(ws, r, "妻の時短中の月収", 250000, "円/月"); r += 1
    r += 1

    # ── 運用パラメータ ──
    section_header(ws, r, "📈  運用パラメータ"); r += 1
    R["return_rate"] = r
    input_row(ws, r, "想定年率リターン", 0.05, "%", "長期株式の標準仮定。控えめに設定推奨"); r += 1
    R["inflation_rate"] = r
    input_row(ws, r, "インフレ率", 0.02, "%", "生活費の年間上昇率"); r += 1
    R["lifespan"] = r
    input_row(ws, r, "生涯年数（何歳まで）", 90, "歳", "FIRE成功の基準年齢"); r += 1

    # 書式設定
    for rn in ["return_rate", "inflation_rate", "income_growth"]:
        ws.cell(row=R[rn], column=2).number_format = "0.0%"
    for rn in ["income_husband", "income_wife", "income_total", "expense_monthly",
               "fire_expense_monthly", "fire_expense_annual",
               "asset_cash", "asset_stocks", "asset_ideco", "asset_total",
               "leave_husband_income", "leave_wife_income", "part_wife_income"]:
        ws.cell(row=R[rn], column=2).number_format = "#,##0"

    return ws


# ══════════════════════════════════════════════
# ② 計算シート
# ══════════════════════════════════════════════
def build_calc_sheet(wb):
    ws = wb.create_sheet("②計算")

    headers = ["経過年", "夫年齢", "妻年齢",
               "夫月収", "妻月収", "世帯月収",
               "月間支出", "月間積立",
               "年間リターン", "年末総資産",
               "FIRE目標額", "FIRE達成"]
    widths  = [8, 8, 8, 14, 14, 14, 12, 12, 14, 16, 16, 10]

    for i, (h, w) in enumerate(zip(headers, widths), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
        c = ws.cell(row=1, column=i, value=h)
        c.font = BOLD; c.fill = GRAY
        c.alignment = Alignment(horizontal="center")
        c.border = thin()

    I = "'①入力'"  # shorthand for cross-sheet reference

    def ref(name):
        return f"{I}!$B${R[name]}"

    for year in range(61):      # 0〜60年
        row = year + 2

        # A: 経過年
        ws.cell(row=row, column=1, value=year)

        # B: 夫年齢 = start + year
        ws.cell(row=row, column=2, value=f"={ref('age_husband')}+{year}")

        # C: 妻年齢
        ws.cell(row=row, column=3, value=f"={ref('age_wife')}+{year}")

        # D: 夫月収
        # IF(夫育休 > 0 かつ year <= ROUNDUP(夫育休/12, 0)) → 夫育休収入
        # ELSE → 夫通常収入 * (1+成長率)^year
        d_formula = (
            f"=IF(AND({ref('leave_husband_months')}>0,"
            f"{year}<=ROUNDUP({ref('leave_husband_months')}/12,0)),"
            f"{ref('leave_husband_income')},"
            f"ROUND({ref('income_husband')}*(1+{ref('income_growth')})^{year},0))"
        )
        ws.cell(row=row, column=4, value=d_formula)

        # E: 妻月収
        # IF(妻育休 > 0 かつ year <= ROUNDUP(妻育休/12, 0)) → 妻育休収入
        # ELIF(妻時短 > 0 かつ year <= ROUNDUP((妻育休+妻時短)/12, 0)) → 妻時短収入
        # ELSE → 妻通常収入 * (1+成長率)^year
        e_formula = (
            f"=IF(AND({ref('leave_wife_months')}>0,"
            f"{year}<=ROUNDUP({ref('leave_wife_months')}/12,0)),"
            f"{ref('leave_wife_income')},"
            f"IF(AND({ref('part_wife_months')}>0,"
            f"{year}<=ROUNDUP(({ref('leave_wife_months')}+{ref('part_wife_months')})/12,0)),"
            f"{ref('part_wife_income')},"
            f"ROUND({ref('income_wife')}*(1+{ref('income_growth')})^{year},0)))"
        )
        ws.cell(row=row, column=5, value=e_formula)

        # F: 世帯月収 = D + E
        ws.cell(row=row, column=6, value=f"=D{row}+E{row}")

        # G: 月間支出（インフレ考慮）
        ws.cell(row=row, column=7,
                value=f"=ROUND({ref('expense_monthly')}*(1+{ref('inflation_rate')})^{year},0)")

        # H: 月間積立 = 世帯月収 - 月間支出
        ws.cell(row=row, column=8, value=f"=F{row}-G{row}")

        # I: 年間リターン（前年末資産 × 年率リターン）
        if year == 0:
            ws.cell(row=row, column=9, value=0)
        else:
            ws.cell(row=row, column=9, value=f"=J{row-1}*{ref('return_rate')}")

        # J: 年末総資産
        if year == 0:
            ws.cell(row=row, column=10, value=f"={ref('asset_total')}")
        else:
            ws.cell(row=row, column=10, value=f"=J{row-1}+H{row}*12+I{row}")

        # K: FIRE目標額（FIRE後年間支出 × 25 × インフレ考慮）
        ws.cell(row=row, column=11,
                value=f"={ref('fire_expense_annual')}*25*(1+{ref('inflation_rate')})^{year}")

        # L: FIRE達成フラグ（初めてJ >= K になった年に 1）
        if year == 0:
            ws.cell(row=row, column=12,
                    value=f"=IF(J{row}>=K{row},1,0)")
        else:
            ws.cell(row=row, column=12,
                    value=f"=IF(AND(J{row}>=K{row},L{row-1}=0),1,0)")

        # 数値書式
        for col in [4, 5, 6, 7, 8, 9, 10, 11]:
            ws.cell(row=row, column=col).number_format = "#,##0"

    # ── N/O 列: FIRE集計 ──
    cw(ws, "N", 26); cw(ws, "O", 18); cw(ws, "P", 8)

    def summary_row(ws, row, label, formula, fmt="", unit=""):
        c = ws.cell(row=row, column=14, value=label)
        c.font = BOLD
        v = ws.cell(row=row, column=15, value=formula)
        v.fill = GREEN; v.border = thin()
        if fmt:
            v.number_format = fmt
        ws.cell(row=row, column=16, value=unit)

    ws.cell(row=1, column=14, value="【FIRE集計】").font = BOLD

    # O2: FIRE到達経過年
    # L1=ヘッダー, L2=year0, ..., L(N+2)=yearN → MATCH位置-2 = 経過年
    summary_row(ws, 2, "FIRE到達経過年",
                '=IFERROR(MATCH(1,L:L,0)-2,"到達なし（60年以内）")', unit="年後")
    # O3: FIRE時夫年齢
    summary_row(ws, 3, "FIRE時 夫年齢",
                f"=IFERROR({ref('age_husband')}+O2,\"-\")", unit="歳")
    # O4: FIRE時妻年齢
    summary_row(ws, 4, "FIRE時 妻年齢",
                f"=IFERROR({ref('age_wife')}+O2,\"-\")", unit="歳")
    # O5: FIRE時総資産
    summary_row(ws, 5, "FIRE時 総資産",
                "=IFERROR(INDEX(J:J,O2+2),\"-\")", "#,##0", "円")
    # O6: 資産倍率（年間支出の何倍か = MCテーブルの軸）
    # 正しくは O5 / fire_expense_annual（×25 は不要）
    summary_row(ws, 6, "FIRE時 資産倍率",
                f"=IFERROR(O5/{ref('fire_expense_annual')},\"-\")", "0.0", "倍")
    # O7: FIRE後年数
    summary_row(ws, 7, "FIRE後の残り年数",
                f"=IFERROR({ref('lifespan')}-O3,\"-\")", unit="年")

    return ws


# ══════════════════════════════════════════════
# ③ FIRE結果シート
# ══════════════════════════════════════════════
def build_result_sheet(wb):
    ws = wb.create_sheet("③FIRE結果")
    cw(ws, "A", 32); cw(ws, "B", 22); cw(ws, "C", 10); cw(ws, "D", 30)

    ws.row_dimensions[1].height = 35
    t = ws.cell(row=1, column=1, value="🔥 FIREシミュレーション結果")
    t.font = BOLD_XL; ws.merge_cells("A1:D1")
    note = ws.cell(row=2, column=1, value="① 入力シートを変更すると自動更新されます。")
    note.font = SMALL; ws.merge_cells("A2:D2")

    r = 4

    def result_row(ws, row, label, formula, unit="", fmt=""):
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=12)
        c = ws.cell(row=row, column=2, value=formula)
        c.fill = GREEN; c.font = Font(bold=True, size=14)
        c.border = thin(); c.alignment = Alignment(horizontal="center")
        if fmt:
            c.number_format = fmt
        ws.cell(row=row, column=3, value=unit).font = Font(size=12)
        ws.row_dimensions[row].height = 26

    # ── FIRE到達予測 ──
    section_header(ws, r, "📅  FIRE到達予測"); r += 1
    result_row(ws, r, "推計FIRE到達年齢（夫）",
               '=IFERROR(②計算!O3&" 歳","60年以内にFIREできません")'); r += 1
    result_row(ws, r, "推計FIRE到達年齢（妻）",
               '=IFERROR(②計算!O4&" 歳","-")'); r += 1
    result_row(ws, r, "FIREまでの年数",
               '=IFERROR(②計算!O2&" 年後","-")'); r += 1
    result_row(ws, r, "FIRE時の推計総資産",
               "=IFERROR(②計算!O5,\"-\")", "円", "#,##0"); r += 1
    result_row(ws, r, "FIRE時の資産倍率（4%ルール）",
               '=IFERROR(ROUND(②計算!O6,1)&" 倍","-")'); r += 1
    r += 1

    # ── MC成功確率 ──
    section_header(ws, r, "📊  FIRE成功確率（モンテカルロ近似）"); r += 1

    def sub_row(ws, row, label, formula, fill=SKYBLUE):
        ws.cell(row=row, column=1, value=label)
        c = ws.cell(row=row, column=2, value=formula)
        c.fill = fill; c.border = thin()
        c.alignment = Alignment(horizontal="center")

    sub_row(ws, r, "参照する資産倍率（5倍刻みに切上）",
            "=IFERROR(CEILING(②計算!O6,5),\"-\")"); r += 1
    sub_row(ws, r, "FIRE後の残り年数（5年刻みに切捨）",
            "=IFERROR(FLOOR(②計算!O7,5),\"-\")"); r += 1

    # 成功確率
    ws.cell(row=r, column=1, value="FIRE成功確率（参考値）").font = Font(bold=True, size=12)
    prob = ws.cell(row=r, column=2, value=(
        "=IFERROR("
        "INDEX('④MCテーブル'!$C$3:$J$9,"
        "MATCH(CEILING(②計算!O6,5),'④MCテーブル'!$B$3:$B$9,0),"
        "MATCH(FLOOR(②計算!O7,5),'④MCテーブル'!$C$2:$J$2,0)"
        ")&\"%\","
        "\"計算不可\")"
    ))
    prob.fill = GREEN; prob.font = Font(bold=True, size=14)
    prob.border = thin(); prob.alignment = Alignment(horizontal="center")
    ws.row_dimensions[r].height = 26
    ws.cell(row=r, column=3, value="80%以上が目安").font = SMALL
    r += 1; r += 1

    # ── 判定 ──
    section_header(ws, r, "💬  判定"); r += 1
    msg = ws.cell(row=r, column=1, value=(
        '=IFERROR(IF(②計算!O6>=25,"✅ FIRE目標達成圏内（資産倍率25倍以上）",'
        '"⚠️ 資産倍率25倍未満。支出を減らすか、資産を増やしましょう"),'
        '"入力値を確認してください")'
    ))
    msg.font = Font(size=12, bold=True); ws.merge_cells(f"A{r}:D{r}"); r += 1; r += 1

    # ── 注意事項 ──
    section_header(ws, r, "📝  注意事項"); r += 1
    notes = [
        "・成功確率は事前計算済みのMCテーブル（2,000回）から近似値を参照しています（精度 ±5%程度）。",
        "・育休・時短の影響は年単位で近似しています（月単位の細かな差は ±1年程度の誤差）。",
        "・年金収入は含みません（保守的な試算です）。",
        "・投資の成果を保証するものではありません。本ツールは情報提供・参考目的のみです。",
        "・投資・退職の判断はご自身の責任で行ってください。",
    ]
    for n in notes:
        c = ws.cell(row=r, column=1, value=n)
        c.font = SMALL; ws.merge_cells(f"A{r}:D{r}"); r += 1

    return ws


# ══════════════════════════════════════════════
# ④ MCテーブルシート
# ══════════════════════════════════════════════
def build_mc_sheet(wb, csv_path):
    ws = wb.create_sheet("④MCテーブル")
    cw(ws, "A", 4); cw(ws, "B", 22)

    ws.cell(row=1, column=1,
            value="MC成功確率テーブル（変更不要）").font = BOLD
    ws.merge_cells("A1:J1")

    with open(csv_path, encoding="utf-8-sig") as f:
        rows = list(csv.reader(f))

    # ヘッダー行（FIRE後年数）
    years = [int(x) for x in rows[0][1:]]
    ws.cell(row=2, column=2, value="資産倍率 \\ FIRE後年数（年）").font = BOLD
    ws.cell(row=2, column=2).fill = GRAY
    for ci, yr in enumerate(years, 3):
        c = ws.cell(row=2, column=ci, value=yr)
        c.font = BOLD; c.fill = GRAY
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(ci)].width = 9

    # データ
    for ri, row_data in enumerate(rows[1:], 3):
        mult = int(row_data[0])
        lbl = ws.cell(row=ri, column=2, value=mult)
        lbl.font = BOLD; lbl.fill = GRAY
        lbl.alignment = Alignment(horizontal="center")
        for ci, val in enumerate(row_data[1:], 3):
            c = ws.cell(row=ri, column=ci, value=float(val))
            c.number_format = "0.0"
            c.alignment = Alignment(horizontal="center")

    src = ws.cell(row=12, column=1,
        value="出典: scripts/generate_mc_table.py（μ=5%, σ=15%, インフレ2%, N=2,000回）")
    src.font = SMALL; ws.merge_cells("A12:J12")
    return ws


# ══════════════════════════════════════════════
# ⑤ 使い方ガイドシート
# ══════════════════════════════════════════════
def build_guide_sheet(wb):
    ws = wb.create_sheet("⑤使い方ガイド")
    cw(ws, "A", 72)

    lines = [
        ("🔥 共働き・子育て世帯のFIREシミュレーター　使い方ガイド", BOLD_XL, 32),
        ("", None, None),
        ("【基本的な使い方】", BOLD, None),
        ("1. 「①入力」シートの黄色いセルに数値を入力してください。", None, None),
        ("2. 「③FIRE結果」シートでFIRE到達年齢・成功確率を確認してください。", None, None),
        ("3. 数値を変えると自動的に結果が更新されます。", None, None),
        ("", None, None),
        ("【各シートの役割】", BOLD, None),
        ("  ① 入力       → 年齢・収入・支出・資産・育休設定などを入力", None, None),
        ("  ② 計算       → 年次シミュレーション表（触らなくてOK）", None, None),
        ("  ③ FIRE結果   → FIRE到達年齢・成功確率・判定メッセージ", None, None),
        ("  ④ MCテーブル → 成功確率の参照テーブル（触らなくてOK）", None, None),
        ("  ⑤ 使い方ガイド → このページ", None, None),
        ("", None, None),
        ("【育休・時短の設定方法】", BOLD, None),
        ("  育休なし        → 夫の育休期間 = 0、妻の育休期間 = 0 に設定", None, None),
        ("  妻が育休12ヶ月  → 妻の育休期間 = 12、妻の育休中の月収 = 育休給付金の額", None, None),
        ("  妻が時短3年     → 妻の時短勤務期間 = 36、妻の時短中の月収 = 時短中の手取り", None, None),
        ("  夫が育休1ヶ月   → 夫の育休期間 = 1、夫の育休中の月収 = 育休給付金の額", None, None),
        ("", None, None),
        ("【育休給付金の目安（2026年現在）】", BOLD, None),
        ("  産後パパ育休（最初28日）: 給与の80%", None, None),
        ("  育休給付金（最初の180日）: 給与の67%", None, None),
        ("  育休給付金（181日〜）    : 給与の50%", None, None),
        ("  ※ 上限額あり・社会保険料は免除（手取りは給付金額より高い）", None, None),
        ("", None, None),
        ("【FIRE成功確率の見方】", BOLD, None),
        ("  80%以上 → 十分安全（1,000人中800人が90歳まで資産維持できる計算）", None, None),
        ("  60〜80% → やや注意（保守的な計画を立てましょう）", None, None),
        ("  60%未満 → 資産倍率を高めることを推奨します", None, None),
        ("  ※ 成功確率はMC近似値（±5%程度の誤差あり）", None, None),
        ("", None, None),
        ("【よくある質問】", BOLD, None),
        ("Q. 結果が「60年以内にFIREできません」と表示される", None, None),
        ("A. 月間積立がマイナス（支出 > 収入）か、FIRE目標額に到達できない状態です。", None, None),
        ("   月間支出を減らすか、FIRE後生活費を下げるか、運用リターンを見直してみてください。", None, None),
        ("", None, None),
        ("Q. 年金はどう考えればいい？", None, None),
        ("A. このシミュレーターには年金を含みません（保守的な試算）。", None, None),
        ("   実際の年金受給見込み額（ねんきんネット）をFIRE後生活費から差し引いて入力してください。", None, None),
        ("", None, None),
        ("Q. iDeCoの引き出しはどう扱う？", None, None),
        ("A. iDeCo残高は現在の資産に含めています。引き出し可能年齢（60歳）の考慮は省略しています。", None, None),
        ("", None, None),
        ("【免責事項】", BOLD, None),
        ("本ツールは教育・参考目的のみです。計算結果は将来の成果を保証するものではありません。", None, None),
        ("投資・退職の判断はご自身の責任で行ってください。", None, None),
    ]

    for ri, (text, font, height) in enumerate(lines, 1):
        c = ws.cell(row=ri, column=1, value=text)
        if font:
            c.font = font
        if height:
            ws.row_dimensions[ri].height = height

    return ws


# ══════════════════════════════════════════════
# メイン
# ══════════════════════════════════════════════
def main():
    os.makedirs("dist", exist_ok=True)
    csv_path = "data/mc_table.csv"

    wb = Workbook()
    wb.remove(wb.active)  # デフォルトシートを削除

    print("①入力シートを作成中...")
    build_input_sheet(wb)
    print(f"  行番号マップ: {R}")

    print("②計算シートを作成中...")
    build_calc_sheet(wb)

    print("③FIRE結果シートを作成中...")
    build_result_sheet(wb)

    print("④MCテーブルシートを作成中...")
    build_mc_sheet(wb, csv_path)

    print("⑤使い方ガイドシートを作成中...")
    build_guide_sheet(wb)

    wb.active = wb["③FIRE結果"]

    out_path = "dist/fire_simulator.xlsx"
    wb.save(out_path)

    print(f"\n生成完了: {out_path}")
    print("\nGoogle スプレッドシートで開く手順:")
    print("  1. drive.google.com を開く")
    print("  2. 「新規」→「ファイルのアップロード」で fire_simulator.xlsx をアップロード")
    print("  3. ファイルを右クリック →「Googleスプレッドシートで開く」")
    print("  4. 「コピーを作成」用の共有URLを取得 → note 有料記事に記載")


if __name__ == "__main__":
    main()
